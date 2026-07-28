import math, io, os, urllib.request, json, time
from PIL import Image, ImageDraw
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Supabase ──────────────────────────────────────────────────
SUPABASE_URL = 'https://limdyowwnlleyyswwkeo.supabase.co'
ANON_KEY = ('eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9'
            '.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImxpbWR5b3d3bmxsZXl5c3d3a2VvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg3NjQyMzYsImV4cCI6MjA5NDM0MDIzNn0'
            '.KJWj-xyBNAg2APrXdOl1ku5lW64wwtV2BWmkIB54080')

# ── Street labels (east end of E-W streets, N→S) ──────────────
STREET_LABELS = [
    (29.35634, -81.06980, 'San Jose Drive'),
    (29.35469, -81.06902, 'Wisteria Drive'),
    (29.35467, -81.06972, 'Juniper Drive'),
    (29.35416, -81.06952, 'Camellia Drive'),
    (29.35330, -81.06835, 'Berkley Road'),
    (29.35246, -81.06795, 'Briggs Drive'),
    (29.35186, -81.06981, 'Raymonde Circle'),
    (29.35100, -81.06725, 'Sunrise Avenue'),
    (29.34985, -81.06998, 'Morning Star Avenue'),
    (29.34928, -81.06644, 'Ocean Breeze Circle'),
    (29.34919, -81.06994, 'Valhalla Avenue'),
    (29.34855, -81.06680, 'Starlight Drive'),
    (29.34851, -81.06990, 'Buckingham Drive'),
    (29.34675, -81.06614, 'Watchtower Drive'),
    (29.34628, -81.06500, 'Spanish Waters Drive'),
    (29.34617, -81.06586, 'Marden Drive'),
    (29.34583, -81.06988, 'Saint Johns Place'),
    (29.34502, -81.06439, 'Kathy Drive'),
    (29.34451, -81.06546, 'Longfellow Circle'),
    (29.34371, -81.06486, 'Ocean Crest Drive'),
    (29.34303, -81.06345, 'Sunset Boulevard'),
    (29.34247, -81.06319, 'Silk Oaks Drive'),
    (29.34126, -81.06261, 'Aqua Vista Drive'),
    (29.34050, -81.06227, 'Sandra Drive'),
    (29.33877, -81.06147, 'Buttenheim Drive'),
    (29.33800, -81.06113, 'Rivocean Drive'),
    (29.33727, -81.06079, 'Seabreeze Drive'),
    (29.33653, -81.06045, 'Seacrest Drive'),
    (29.33578, -81.06011, 'Holland Road'),
    (29.33505, -81.06062, 'Concord Drive'),
    (29.33447, -81.06242, 'Sunrise Cove Circle'),
    (29.33439, -81.05947, 'Town and Country Lane'),
    (29.33404, -81.05933, 'Roberta Road'),
    (29.33334, -81.05914, 'Laurie Drive'),
    (29.33229, -81.05990, 'Margaret Road'),
    (29.33100, -81.05954, 'Ocean Shore Drive'),
    (29.33075, -81.05950, 'River Shore Drive'),
    (29.33026, -81.05744, 'Ocean Edge Drive'),
    (29.32978, -81.05919, 'Seaside Drive'),
]

def closest_street(lat, lng):
    best, best_d = '', float('inf')
    for slat, slng, name in STREET_LABELS:
        # Weight lat slightly (1 deg lat ≈ 111km, 1 deg lng ≈ 99km here)
        d = math.sqrt(((lat - slat) * 111)**2 + ((lng - slng) * 99)**2)
        if d < best_d:
            best_d = d
            best = name
    return best

def lat_lng_to_tile(lat, lng, z):
    n = 2 ** z
    x = int((lng + 180.0) / 360.0 * n)
    lat_r = math.radians(lat)
    y = int((1.0 - math.asinh(math.tan(lat_r)) / math.pi) / 2.0 * n)
    return x, y

def px_offset_in_tile(lat, lng, tx, ty, z):
    n = 2 ** z
    tile_lng0 = tx / n * 360.0 - 180.0
    tile_lng1 = (tx + 1) / n * 360.0 - 180.0
    px = (lng - tile_lng0) / (tile_lng1 - tile_lng0) * 256

    def merc(la): return math.log(math.tan(math.pi/4 + math.radians(la)/2))
    top_m = merc(math.degrees(math.atan(math.sinh(math.pi*(1 - 2*ty/n)))))
    bot_m = merc(math.degrees(math.atan(math.sinh(math.pi*(1 - 2*(ty+1)/n)))))
    py = (top_m - merc(lat)) / (top_m - bot_m) * 256
    return px, py

HEADERS_ESRI = {'User-Agent': 'turtle-tracker-report/1.0'}

def fetch_tile(z, ty, tx):
    url = f'https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{ty}/{tx}'
    req = urllib.request.Request(url, headers=HEADERS_ESRI)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=15) as r:
                return Image.open(io.BytesIO(r.read())).convert('RGB')
        except Exception as e:
            if attempt == 2: raise
            time.sleep(1)

def make_map_image(lat, lng, zoom=17, crop_px=220):
    tx, ty = lat_lng_to_tile(lat, lng, zoom)
    composite = Image.new('RGB', (768, 768), (30, 30, 30))
    for dy in (-1, 0, 1):
        for dx in (-1, 0, 1):
            try:
                tile = fetch_tile(zoom, ty + dy, tx + dx)
                composite.paste(tile, ((dx+1)*256, (dy+1)*256))
            except:
                pass
    px, py = px_offset_in_tile(lat, lng, tx, ty, zoom)
    cx = int(256 + px)
    cy = int(256 + py)
    # Draw marker
    draw = ImageDraw.Draw(composite)
    r = 9
    draw.ellipse([cx-r, cy-r, cx+r, cy+r], fill=(220, 40, 40), outline='white', width=2)
    draw.ellipse([cx-3, cy-3, cx+3, cy+3], fill='white')
    # Crop square around marker
    half = crop_px
    box = (max(0, cx-half), max(0, cy-half), min(768, cx+half), min(768, cy+half))
    img = composite.crop(box).resize((180, 180), Image.LANCZOS)
    return img

def fetch_photo(url, max_dim=150):
    req = urllib.request.Request(url, headers={'User-Agent': 'turtle-tracker-report/1.0'})
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            img = Image.open(io.BytesIO(r.read())).convert('RGB')
        # Crop to square from center top (nest stakes are usually near top center)
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        img = img.crop((left, 0, left + side, side))
        img.thumbnail((max_dim, max_dim), Image.LANCZOS)
        return img
    except:
        return None

def img_to_xl(img, tmp_path):
    img.save(tmp_path, 'PNG')
    return XLImage(tmp_path)

def fmt_date(s):
    if not s: return ''
    parts = s.split('-')
    if len(parts) == 3:
        m = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        try: return f"{m[int(parts[1])-1]} {int(parts[2])}, {parts[0]}"
        except: pass
    return s

# ── Fetch nests ───────────────────────────────────────────────
url = f'{SUPABASE_URL}/rest/v1/turtle_nests?select=*&order=date_found.asc'
req = urllib.request.Request(url, headers={'apikey': ANON_KEY, 'Authorization': f'Bearer {ANON_KEY}'})
with urllib.request.urlopen(req) as r:
    nests = json.loads(r.read())
print(f"Fetched {len(nests)} nests")

# ── Build workbook ────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Nest Log'

# Column widths (in Excel units)
ws.column_dimensions['A'].width = 14   # Nest #
ws.column_dimensions['B'].width = 18   # Photo
ws.column_dimensions['C'].width = 24   # Map
ws.column_dimensions['D'].width = 24   # Closest Street
ws.column_dimensions['E'].width = 14   # Found Date
ws.column_dimensions['F'].width = 22   # Hatch Window

# Header row
HEADER_FILL = PatternFill('solid', start_color='1A3A5C')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADERS = ['Nest #', 'Nest Photo', 'Location Map', 'Nearest Street', 'Date Found', 'Hatch Window']
for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 24

# Data rows
ALT_FILL  = PatternFill('solid', start_color='EAF2FB')
BASE_FILL = PatternFill('solid', start_color='FFFFFF')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TMP = '/tmp/turtle_img_cache'
os.makedirs(TMP, exist_ok=True)

IMG_ROW_HEIGHT = 140  # points — roughly 140px
IMG_H_PT = 130        # image display height in points (slightly smaller than row)
IMG_W_PT_PHOTO = 96
IMG_W_PT_MAP = 130

for i, nest in enumerate(nests):
    row = i + 2
    fill = ALT_FILL if i % 2 else BASE_FILL
    street = closest_street(nest['latitude'], nest['longitude'])
    is_fc = nest['species'] == 'false_crawl'

    hatch = ('False Crawl' if is_fc
             else f"{fmt_date(nest['hatch_date_min'])} – {fmt_date(nest['hatch_date_max'])}")

    # Text cells
    def tc(col, val, bold=False, center=False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Arial', bold=bold, size=10)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center' if center else 'left',
                                   vertical='center', wrap_text=True)

    tc(1, nest['nest_number'], bold=True, center=True)
    # B and C will have images (leave blank)
    ws.cell(row=row, column=2).fill = fill
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3).fill = fill
    ws.cell(row=row, column=3).border = BORDER
    tc(4, street)
    tc(5, fmt_date(nest['date_found']), center=True)
    tc(6, hatch)

    ws.row_dimensions[row].height = IMG_ROW_HEIGHT

    # ── Nest photo ──────────────────────────────────────────
    if nest.get('photo_url'):
        print(f"  Photo for {nest['nest_number']}…")
        photo = fetch_photo(nest['photo_url'])
        if photo:
            tmp = f'{TMP}/photo_{i}.png'
            xl_img = img_to_xl(photo, tmp)
            xl_img.width  = IMG_W_PT_PHOTO
            xl_img.height = IMG_H_PT
            cell_addr = f'B{row}'
            ws.add_image(xl_img, cell_addr)

    # ── Map ─────────────────────────────────────────────────
    print(f"  Map for {nest['nest_number']}…")
    try:
        map_img = make_map_image(nest['latitude'], nest['longitude'])
        tmp = f'{TMP}/map_{i}.png'
        xl_map = img_to_xl(map_img, tmp)
        xl_map.width  = IMG_W_PT_MAP
        xl_map.height = IMG_H_PT
        ws.add_image(xl_map, f'C{row}')
    except Exception as e:
        print(f"    Map failed: {e}")

    time.sleep(0.3)  # be polite to tile server

# Freeze header
ws.freeze_panes = 'A2'

OUT = '/Users/PhillipHurst/Desktop/Turtle_Nest_Log.xlsx'
wb.save(OUT)
print(f"\nSaved → {OUT}")
